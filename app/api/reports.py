from io import BytesIO
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, or_, func, case
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from ..database import get_db
from ..models import StockMovement, Partner, Item, Payment, User
from ..schemas import (
    Page, TransactionOut, StatTotal, TransactionPeriod, TransactionPartnerSubtotal,
    TransactionItemSubtotal, AgingReport, AgingBucket,
)
from ..deps import require_permission
from ..timeutil import now_business, today
from ..services import paginate

router = APIRouter(prefix="/api/reports", tags=["reports"])


@router.get("/aging", response_model=AgingReport)
def aging_report(
    kind: str = Query("AR", pattern="^(AR|AP)$", description="AR(미수)/AP(미지급)"),
    as_of: str | None = Query(None, description="기준일 YYYY-MM-DD (미지정 시 오늘)"),
    partner_id: int | None = Query(None, description="특정 거래처만"),
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("payment:read")),
):
    """미수(AR)/미지급(AP) 연령분석. 청구(이동)에 결제·반품을 FIFO(오래된 청구부터)로
    상계하고, 남은 미결제 금액을 청구 경과일 구간(0-30/31-60/61-90/90+)으로 분류한다."""
    try:
        # 기준일 미지정이면 '사업장 기준 오늘'(앱 프로세스의 로컬 시각이 아니라)
        as_of_dt = (
            datetime.strptime(as_of, "%Y-%m-%d") if as_of
            else datetime.combine(today(), datetime.min.time())
        )
    except ValueError:
        raise HTTPException(status_code=400, detail="기준일 형식이 올바르지 않습니다 (YYYY-MM-DD)")
    as_of_dt = as_of_dt.replace(hour=23, minute=59, second=59)
    mv_type = "OUT" if kind == "AR" else "IN"

    # 청구 이동(양수 수량): (거래처, 일자, 금액=수량×단가+세액)
    charge_rows = db.execute(
        select(StockMovement.partner_id, StockMovement.created_at,
               StockMovement.quantity, StockMovement.unit_price, StockMovement.tax_amount)
        .where(StockMovement.movement_type == mv_type, StockMovement.quantity > 0,
               StockMovement.partner_id.isnot(None))
    ).all()
    # 반품 크레딧(음수 수량 → 감액) 거래처별 합
    ret_rows = db.execute(
        select(StockMovement.partner_id,
               func.coalesce(func.sum(
                   StockMovement.quantity * StockMovement.unit_price + StockMovement.tax_amount), 0))
        .where(StockMovement.movement_type == mv_type, StockMovement.quantity < 0,
               StockMovement.partner_id.isnot(None))
        .group_by(StockMovement.partner_id)
    ).all()
    # 결제 크레딧 거래처별 합
    pay_rows = db.execute(
        select(Payment.partner_id, func.coalesce(func.sum(Payment.amount), 0))
        .where(Payment.kind == kind).group_by(Payment.partner_id)
    ).all()

    credit: dict[int, int] = {}
    for pid, s in ret_rows:
        credit[pid] = credit.get(pid, 0) - int(s)   # 반품 청구합(음수)의 절대값을 크레딧으로
    for pid, s in pay_rows:
        credit[pid] = credit.get(pid, 0) + int(s)

    charges: dict[int, list] = {}
    for pid, created, qty, price, tax in charge_rows:
        if partner_id and pid != partner_id:
            continue
        charges.setdefault(pid, []).append((created, qty * price + int(tax)))

    names = dict(db.execute(select(Partner.id, Partner.name)).all())
    buckets, grand = [], 0
    for pid, lst in charges.items():
        lst.sort(key=lambda x: x[0])   # 오래된 청구부터
        cr = credit.get(pid, 0)
        b = {"b_0_30": 0, "b_31_60": 0, "b_61_90": 0, "b_over_90": 0}
        total = 0
        for created, amt in lst:
            unpaid = amt
            if cr > 0:
                applied = min(cr, unpaid)
                unpaid -= applied
                cr -= applied
            if unpaid <= 0:
                continue
            age = (as_of_dt - created).days
            key = "b_0_30" if age <= 30 else "b_31_60" if age <= 60 else "b_61_90" if age <= 90 else "b_over_90"
            b[key] += unpaid
            total += unpaid
        if total <= 0:
            continue
        buckets.append(AgingBucket(partner_id=pid, partner_name=names.get(pid, ""), total=total, **b))
        grand += total
    buckets.sort(key=lambda x: x.total, reverse=True)
    return AgingReport(kind=kind, as_of=as_of_dt.strftime("%Y-%m-%d"), buckets=buckets, total=grand)

# 소계 시트(집계행) 정렬 화이트리스트. 미허용 값은 export에서 400으로 거부한다.
# 주의: 이 키(purchase_amount/sales_amount/... 매입·매출 분리 집계)는 개별 이동행 목록
# 정렬 키인 아래 _SORT_COLUMNS(amount/quantity 등 행 단위 값)와 다른 축이다. 혼동 금지.
_PARTNER_SORT_KEYS = {
    "partner_name": lambda b: b["name"],
    "purchase_amount": lambda b: b["purchase_amount"],
    "sales_amount": lambda b: b["sales_amount"],
    "purchase_count": lambda b: b["purchase_count"],
    "sales_count": lambda b: b["sales_count"],
}
_ITEM_SORT_KEYS = {
    "item_name": lambda b: b["name"],
    "purchase_amount": lambda b: b["purchase_amount"],
    "sales_amount": lambda b: b["sales_amount"],
    "purchase_count": lambda b: b["purchase_count"],
    "sales_count": lambda b: b["sales_count"],
}

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _parse_date(value: str) -> datetime:
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="날짜 형식이 올바르지 않습니다 (YYYY-MM-DD)")


def _validate_order(value: str) -> None:
    """정렬 방향은 asc/desc 만 허용한다. 오타를 조용히 desc로 흡수하지 않는다."""
    if value not in ("asc", "desc"):
        raise HTTPException(
            status_code=400,
            detail=f"정렬 방향은 asc 또는 desc 여야 합니다: {value}",
        )


def _apply_filters(stmt, q, kind, partner_id, date_from, date_to):
    """매입/매출 공통 필터. IN=매입, OUT=매출이며 ADJUST(재고 조정)는 제외한다."""
    stmt = stmt.where(StockMovement.movement_type.in_(("IN", "OUT")))
    if q:
        stmt = (
            stmt
            .join(Item, StockMovement.item_id == Item.id)
            .outerjoin(Partner, StockMovement.partner_id == Partner.id)
            .where(or_(
                StockMovement.movement_no.like(f"%{q}%"),
                Item.name.like(f"%{q}%"),
                Item.code.like(f"%{q}%"),
                Partner.name.like(f"%{q}%"),
            ))
        )
    if partner_id is not None:
        stmt = stmt.where(StockMovement.partner_id == partner_id)
    if kind == "purchase":
        stmt = stmt.where(StockMovement.movement_type == "IN")
    elif kind == "sales":
        stmt = stmt.where(StockMovement.movement_type == "OUT")
    if date_from:
        stmt = stmt.where(StockMovement.created_at >= _parse_date(date_from))
    if date_to:
        stmt = stmt.where(StockMovement.created_at < _parse_date(date_to) + timedelta(days=1))
    return stmt


def _txn_out(m: StockMovement) -> TransactionOut:
    """입출고 이력을 매입/매출 내역 응답 스키마로 변환한다. (IN=매입, OUT=매출)"""
    return TransactionOut(
        id=m.id,
        movement_no=m.movement_no,
        created_at=m.created_at,
        kind="purchase" if m.movement_type == "IN" else "sales",
        partner_id=m.partner_id,
        partner_name=m.partner.name if m.partner else "",
        item_id=m.item_id,
        item_code=m.item.code,
        item_name=m.item.name,
        quantity=m.quantity,
        unit_price=m.unit_price,
        amount=m.quantity * m.unit_price,
    )


# 매입/매출 내역 목록(개별 이동행) 정렬 허용 컬럼. 거래처/품목명은 조인이 필요해 제외.
# 위 _PARTNER_SORT_KEYS/_ITEM_SORT_KEYS(소계 집계행 정렬)와는 다른 축이다.
_SORT_COLUMNS = {
    "created_at": StockMovement.created_at,
    "movement_no": StockMovement.movement_no,
    "quantity": StockMovement.quantity,
    "unit_price": StockMovement.unit_price,
    "amount": StockMovement.quantity * StockMovement.unit_price,
}


@router.get("/transactions", response_model=Page[TransactionOut])
def list_transactions(
    q: str | None = Query(None, description="번호·품목·거래처 검색"),
    kind: str | None = Query(None, description="purchase/sales 로 필터"),
    partner_id: int | None = Query(None, description="특정 거래처만 조회"),
    date_from: str | None = Query(None, description="시작일 YYYY-MM-DD"),
    date_to: str | None = Query(None, description="종료일 YYYY-MM-DD(포함)"),
    sort: str = Query("created_at", description="정렬 기준(created_at/movement_no/quantity/unit_price/amount)"),
    order: str = Query("desc", description="정렬 방향(asc/desc)"),
    page: int = Query(1, ge=1, description="페이지 번호(1부터)"),
    page_size: int = Query(20, ge=1, le=100, description="페이지당 건수(최대 100)"),
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("stock:read")),
):
    """전사 매입/매출 내역 목록. sort/order 기준으로 정렬한다(기본 최신순)."""
    if sort not in _SORT_COLUMNS:
        raise HTTPException(
            status_code=400,
            detail=f"허용되지 않은 정렬 컬럼입니다: {sort}",
        )
    _validate_order(order)
    col = _SORT_COLUMNS[sort]
    direction = col.asc() if order == "asc" else col.desc()
    stmt = _apply_filters(
        select(StockMovement), q, kind, partner_id, date_from, date_to
    ).order_by(direction, StockMovement.id.desc())
    result = paginate(db, stmt, page, page_size)
    result["items"] = [_txn_out(m) for m in result["items"]]
    return result


@router.get("/transactions/summary", response_model=StatTotal)
def get_transactions_summary(
    q: str | None = Query(None, description="번호·품목·거래처 검색"),
    kind: str | None = Query(None, description="purchase/sales 로 필터"),
    partner_id: int | None = Query(None, description="특정 거래처만 조회"),
    date_from: str | None = Query(None, description="시작일 YYYY-MM-DD"),
    date_to: str | None = Query(None, description="종료일 YYYY-MM-DD(포함)"),
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("stock:read")),
):
    """현재 필터 조건의 매입/매출 합계·건수."""
    amount = StockMovement.quantity * StockMovement.unit_price
    agg = _apply_filters(
        select(
            func.sum(case((StockMovement.movement_type == "IN", amount), else_=0)),
            func.sum(case((StockMovement.movement_type == "OUT", amount), else_=0)),
            func.sum(case((StockMovement.movement_type == "IN", 1), else_=0)),
            func.sum(case((StockMovement.movement_type == "OUT", 1), else_=0)),
        ),
        q, kind, partner_id, date_from, date_to,
    )
    row = db.execute(agg).one()
    return StatTotal(
        purchase_amount=int(row[0] or 0),
        sales_amount=int(row[1] or 0),
        purchase_count=int(row[2] or 0),
        sales_count=int(row[3] or 0),
    )


@router.get("/transactions/periods", response_model=list[TransactionPeriod])
def get_transactions_periods(
    q: str | None = Query(None, description="번호·품목·거래처 검색"),
    kind: str | None = Query(None, description="purchase/sales 로 필터"),
    partner_id: int | None = Query(None, description="특정 거래처만 조회"),
    date_from: str | None = Query(None, description="시작일 YYYY-MM-DD"),
    date_to: str | None = Query(None, description="종료일 YYYY-MM-DD(포함)"),
    group_by: str = Query("month", description="집계 단위(month/day)"),
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("stock:read")),
):
    """현재 필터 조건의 매입/매출을 기간(월/일)별로 집계한다. 최신 기간이 위로 오도록 내림차순."""
    rows = db.execute(
        _apply_filters(
            select(
                StockMovement.created_at,
                StockMovement.movement_type,
                StockMovement.quantity,
                StockMovement.unit_price,
            ),
            q, kind, partner_id, date_from, date_to,
        )
    ).all()
    # DB 이식성(MySQL/SQLite)을 위해 날짜 버킷팅은 SQL이 아닌 파이썬에서 수행한다
    fmt = "%Y-%m-%d" if group_by == "day" else "%Y-%m"
    buckets: dict[str, dict[str, int]] = {}
    for created_at, movement_type, quantity, unit_price in rows:
        key = created_at.strftime(fmt)
        b = buckets.setdefault(key, {
            "purchase_amount": 0, "sales_amount": 0,
            "purchase_count": 0, "sales_count": 0,
        })
        prefix = "purchase" if movement_type == "IN" else "sales"
        b[f"{prefix}_amount"] += quantity * unit_price
        b[f"{prefix}_count"] += 1
    return [
        TransactionPeriod(period=key, **buckets[key])
        for key in sorted(buckets, reverse=True)
    ]


@router.get("/transactions/partners", response_model=list[TransactionPartnerSubtotal])
def get_transactions_partners(
    q: str | None = Query(None, description="번호·품목·거래처 검색"),
    kind: str | None = Query(None, description="purchase/sales 로 필터"),
    partner_id: int | None = Query(None, description="특정 거래처만 조회"),
    date_from: str | None = Query(None, description="시작일 YYYY-MM-DD"),
    date_to: str | None = Query(None, description="종료일 YYYY-MM-DD(포함)"),
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("stock:read")),
):
    """현재 필터 조건의 매입/매출을 거래처별로 집계한다. 합계 금액이 큰 순으로 정렬."""
    rows = db.execute(
        _apply_filters(
            select(
                StockMovement.partner_id,
                StockMovement.movement_type,
                StockMovement.quantity,
                StockMovement.unit_price,
            ),
            q, kind, partner_id, date_from, date_to,
        )
    ).all()
    # _apply_filters가 q 검색 시 Partner를 이미 조인하므로, 이중 조인을 피해 파이썬에서 버킷팅한다
    buckets: dict[int | None, dict[str, int]] = {}
    for pid, movement_type, quantity, unit_price in rows:
        b = buckets.setdefault(pid, {
            "purchase_amount": 0, "sales_amount": 0,
            "purchase_count": 0, "sales_count": 0,
        })
        prefix = "purchase" if movement_type == "IN" else "sales"
        b[f"{prefix}_amount"] += quantity * unit_price
        b[f"{prefix}_count"] += 1
    ids = [pid for pid in buckets if pid is not None]
    name_map = {}
    if ids:
        name_map = dict(db.execute(select(Partner.id, Partner.name).where(Partner.id.in_(ids))).all())
    result = [
        TransactionPartnerSubtotal(
            partner_id=pid,
            partner_name=name_map.get(pid, "미지정") if pid is not None else "미지정",
            **b,
        )
        for pid, b in buckets.items()
    ]
    result.sort(key=lambda r: r.purchase_amount + r.sales_amount, reverse=True)
    return result


@router.get("/transactions/items", response_model=list[TransactionItemSubtotal])
def get_transactions_items(
    q: str | None = Query(None, description="번호·품목·거래처 검색"),
    kind: str | None = Query(None, description="purchase/sales 로 필터"),
    partner_id: int | None = Query(None, description="특정 거래처만 조회"),
    date_from: str | None = Query(None, description="시작일 YYYY-MM-DD"),
    date_to: str | None = Query(None, description="종료일 YYYY-MM-DD(포함)"),
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("stock:read")),
):
    """현재 필터 조건의 매입/매출을 품목별로 집계한다. 합계 금액이 큰 순으로 정렬."""
    rows = db.execute(
        _apply_filters(
            select(
                StockMovement.item_id,
                StockMovement.movement_type,
                StockMovement.quantity,
                StockMovement.unit_price,
            ),
            q, kind, partner_id, date_from, date_to,
        )
    ).all()
    # _apply_filters가 q 검색 시 Item을 이미 조인하므로, 이중 조인을 피해 파이썬에서 버킷팅한다
    buckets: dict[int, dict[str, int]] = {}
    for iid, movement_type, quantity, unit_price in rows:
        b = buckets.setdefault(iid, {
            "purchase_amount": 0, "sales_amount": 0,
            "purchase_count": 0, "sales_count": 0,
        })
        prefix = "purchase" if movement_type == "IN" else "sales"
        b[f"{prefix}_amount"] += quantity * unit_price
        b[f"{prefix}_count"] += 1
    ids = list(buckets)
    name_map: dict[int, tuple[str, str]] = {}
    if ids:
        name_map = {
            iid: (code, name)
            for iid, code, name in db.execute(
                select(Item.id, Item.code, Item.name).where(Item.id.in_(ids))
            ).all()
        }
    result = [
        TransactionItemSubtotal(
            item_id=iid,
            item_code=name_map.get(iid, ("", ""))[0],
            item_name=name_map.get(iid, ("", ""))[1],
            **b,
        )
        for iid, b in buckets.items()
    ]
    result.sort(key=lambda r: r.purchase_amount + r.sales_amount, reverse=True)
    return result


@router.get("/transactions/export")
def export_transactions(
    q: str | None = Query(None, description="번호·품목·거래처 검색"),
    partner_id: int | None = Query(None, description="특정 거래처만 내보내기"),
    kind: str | None = Query(None, description="purchase/sales 로 필터"),
    date_from: str | None = Query(None, description="시작일 YYYY-MM-DD"),
    date_to: str | None = Query(None, description="종료일 YYYY-MM-DD(포함)"),
    item_sort: str | None = Query(None, description="품목별 소계 정렬 컬럼(item_name/purchase_amount/sales_amount/purchase_count/sales_count)"),
    item_order: str = Query("desc", description="품목별 소계 정렬 방향 asc/desc"),
    partner_sort: str | None = Query(None, description="거래처별 소계 정렬 컬럼(partner_name/purchase_amount/sales_amount/purchase_count/sales_count)"),
    partner_order: str = Query("desc", description="거래처별 소계 정렬 방향 asc/desc"),
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("stock:read")),
):
    """매입/매출 내역을 엑셀(.xlsx)로 내보낸다. IN=매입, OUT=매출이며 ADJUST(재고 조정)는 제외."""
    if partner_sort is not None and partner_sort not in _PARTNER_SORT_KEYS:
        raise HTTPException(
            status_code=400,
            detail=f"허용되지 않은 거래처 정렬 컬럼입니다: {partner_sort}",
        )
    if item_sort is not None and item_sort not in _ITEM_SORT_KEYS:
        raise HTTPException(
            status_code=400,
            detail=f"허용되지 않은 품목 정렬 컬럼입니다: {item_sort}",
        )
    _validate_order(partner_order)
    _validate_order(item_order)

    stmt = _apply_filters(
        select(StockMovement), q, kind, partner_id, date_from, date_to
    ).order_by(StockMovement.created_at.asc(), StockMovement.id.asc())

    movements = db.execute(stmt).scalars().all()

    # ---- 엑셀 작성 ----
    wb = Workbook()
    ws = wb.active
    ws.title = "매입매출내역"

    bold = Font(bold=True)
    right = Alignment(horizontal="right")

    headers = ["일자", "번호", "구분", "거래처", "품목코드", "품목명", "수량", "단가", "금액"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        if col >= 7:  # 수량/단가/금액
            cell.alignment = right

    purchase_total = 0
    sales_total = 0
    for m in movements:
        amount = m.quantity * m.unit_price
        if m.movement_type == "IN":
            purchase_total += amount
        else:
            sales_total += amount
        ws.append([
            m.created_at.strftime("%Y-%m-%d"),
            m.movement_no,
            "매입" if m.movement_type == "IN" else "매출",
            m.partner.name if m.partner else "",
            m.item.code,
            m.item.name,
            m.quantity,
            m.unit_price,
            amount,
        ])
        for col in (7, 8, 9):  # 수량/단가/금액 천단위 구분
            ws.cell(row=ws.max_row, column=col).number_format = "#,##0"

    # ---- 합계 (매입/매출 별도) ----
    ws.append([])
    for label, total in (("매입 합계", purchase_total), ("매출 합계", sales_total)):
        ws.append(["", "", label, "", "", "", "", "", total])
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).font = bold
        ws.cell(row=ws.max_row, column=9).number_format = "#,##0"

    # ---- 열 너비 ----
    widths = {"A": 12, "B": 18, "C": 8, "D": 18, "E": 14, "F": 20, "G": 10, "H": 12, "I": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # ---- 거래처별 소계 시트 ----
    buckets: dict[int | None, dict] = {}
    for m in movements:
        b = buckets.setdefault(m.partner_id, {
            "name": m.partner.name if m.partner else "미지정",
            "purchase_amount": 0,
            "sales_amount": 0,
            "purchase_count": 0,
            "sales_count": 0,
        })
        amount = m.quantity * m.unit_price
        if m.movement_type == "IN":
            b["purchase_amount"] += amount
            b["purchase_count"] += 1
        else:
            b["sales_amount"] += amount
            b["sales_count"] += 1

    # 프론트에서 지정한 정렬을 반영. 미지정이면 합계 금액 내림차순(기본값). 미허용 값은 위에서 400으로 거부됨.
    if partner_sort in _PARTNER_SORT_KEYS:
        rows = sorted(
            buckets.values(),
            key=_PARTNER_SORT_KEYS[partner_sort],
            reverse=(partner_order != "asc"),
        )
    else:
        rows = sorted(
            buckets.values(),
            key=lambda b: b["purchase_amount"] + b["sales_amount"],
            reverse=True,
        )

    ws2 = wb.create_sheet("거래처별소계")
    ws2.append(["거래처", "매입 합계", "매출 합계", "매입 건수", "매출 건수"])
    for col in range(1, 6):
        cell = ws2.cell(row=1, column=col)
        cell.font = bold
        if col >= 2:
            cell.alignment = right
    for b in rows:
        ws2.append([
            b["name"],
            b["purchase_amount"],
            b["sales_amount"],
            b["purchase_count"],
            b["sales_count"],
        ])
        for col in range(2, 6):
            ws2.cell(row=ws2.max_row, column=col).number_format = "#,##0"
    for col, width in {"A": 22, "B": 14, "C": 14, "D": 10, "E": 10}.items():
        ws2.column_dimensions[col].width = width

    # ---- 품목별 소계 시트 ----
    item_buckets: dict[int, dict] = {}
    for m in movements:
        b = item_buckets.setdefault(m.item_id, {
            "code": m.item.code,
            "name": m.item.name,
            "purchase_amount": 0,
            "sales_amount": 0,
            "purchase_count": 0,
            "sales_count": 0,
        })
        amount = m.quantity * m.unit_price
        if m.movement_type == "IN":
            b["purchase_amount"] += amount
            b["purchase_count"] += 1
        else:
            b["sales_amount"] += amount
            b["sales_count"] += 1

    # 프론트에서 지정한 정렬을 반영. 미지정이면 합계 금액 내림차순(기본값). 미허용 값은 위에서 400으로 거부됨.
    if item_sort in _ITEM_SORT_KEYS:
        item_rows = sorted(
            item_buckets.values(),
            key=_ITEM_SORT_KEYS[item_sort],
            reverse=(item_order != "asc"),
        )
    else:
        item_rows = sorted(
            item_buckets.values(),
            key=lambda b: b["purchase_amount"] + b["sales_amount"],
            reverse=True,
        )

    ws3 = wb.create_sheet("품목별소계")
    ws3.append(["품목코드", "품목명", "매입 합계", "매출 합계", "매입 건수", "매출 건수"])
    for col in range(1, 7):
        cell = ws3.cell(row=1, column=col)
        cell.font = bold
        if col >= 3:
            cell.alignment = right
    for b in item_rows:
        ws3.append([
            b["code"],
            b["name"],
            b["purchase_amount"],
            b["sales_amount"],
            b["purchase_count"],
            b["sales_count"],
        ])
        for col in range(3, 7):
            ws3.cell(row=ws3.max_row, column=col).number_format = "#,##0"
    for col, width in {"A": 14, "B": 22, "C": 14, "D": 14, "E": 10, "F": 10}.items():
        ws3.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 파일명은 헤더 인코딩 문제를 피하기 위해 ASCII만 사용
    filename = f"sales_purchase_{now_business().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
